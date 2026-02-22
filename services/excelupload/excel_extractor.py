import io
from openpyxl import load_workbook
from typing import List, Dict, Any

def parse_header(sheet) -> Dict[str, str]:
    header_cell = sheet['A1'].value or ''
    header_parts = header_cell.split('-')
    if len(header_parts) < 2:
        raise ValueError('Header format invalid')
    payingAgentAndCompartment = header_parts[0].strip()
    ISINNumber = header_parts[1].strip()
    payingAgentCompartment_parts = payingAgentAndCompartment.split(' ')
    if len(payingAgentCompartment_parts) < 2:
        raise ValueError('Paying agent/compartment format invalid')
    payingAgent = payingAgentCompartment_parts[0]
    compartment = payingAgentCompartment_parts[1]
    return {
        'payingAgent': payingAgent,
        'compartment': compartment,
        'ISINNumber': ISINNumber.strip()
    }

def is_strikethrough(cell) -> bool:
    if cell.font and hasattr(cell.font, 'strike'):
        return cell.font.strike
    return False

def parse_body(sheet) -> List[Dict[str, Any]]:
    data = []
    for row in sheet.iter_rows(min_row=4):
        tradeDate = row[0].value
        counterParty = row[1].value
        notional = row[2].value
        reference = row[3].value
        notional_value = notional
        subscriptionCancelled = False
        if is_strikethrough(row[2]):
            try:
                notional_value = -abs(float(notional))
            except Exception:
                notional_value = notional
            subscriptionCancelled = True
        elif is_strikethrough(row[0]) or is_strikethrough(row[1]) or is_strikethrough(row[3]):
            subscriptionCancelled = True

        if all(v is None for v in [tradeDate, counterParty, notional, reference]):
            continue
        data.append({
            'tradeDate': tradeDate,
            'counterParty': counterParty,
            'notional': notional_value,
            'reference': reference,
            'subscriptionCancelled': subscriptionCancelled
        })
    return data

def process_excel(file_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = []
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        try:
            header = parse_header(sheet)
            body = parse_body(sheet)
            result.append({
                'sheet': sheetname,
                'header': header,
                'body': body
            })
        except Exception as e:
            result.append({
                'sheet': sheetname,
                'error': str(e)
            })
    return result
