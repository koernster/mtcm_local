"""
Excel/CSV-based contact provider — fallback when HubSpot API is unavailable.

Reads contacts from a CSV or Excel file (HubSpot export format) and serves
them through the same interface as the live HubSpot API.

Supports both Spanish and English HubSpot export column headers.
Prefers CSV (more reliable) but also handles .xlsx files.
"""

import os
import csv
import logging
from typing import List, Dict, Optional
from datetime import datetime

logger = logging.getLogger(__name__)

# ── Column mapping: HubSpot export header → internal field name ──────────────
COLUMN_MAP = {
    # Spanish
    "ID de registro": "id",
    "Nombre": "firstname",
    "Apellidos": "lastname",
    "Fecha de creación": "createdate",
    "Correo": "email",
    "Número de teléfono": "phone",
    "Estado del lead": "hs_lead_status",
    "Propietario del contacto": "hubspot_owner_id",
    "Estado del contacto de marketing": "marketing_contact_status",
    # English
    "Record ID": "id",
    "First Name": "firstname",
    "Last Name": "lastname",
    "Create Date": "createdate",
    "Email": "email",
    "Phone Number": "phone",
    "Lead Status": "hs_lead_status",
    "Contact owner": "hubspot_owner_id",
    "Marketing contact status": "marketing_contact_status",
    "Last Activity Date": "last_activity_date",
    "Associated Company": "company",
    "Primary Associated Company ID": "associated_company_id",
    "Additional email addresses": "additional_emails",
}


def _normalise_value(value) -> str:
    """Convert cell value to a JSON-safe string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat() + "Z"
    return str(value).strip()


def _clean_id(raw_id: str) -> str:
    """Normalise a HubSpot record ID (may be scientific notation or float)."""
    if not raw_id:
        return ""
    try:
        return str(int(float(raw_id)))
    except (ValueError, OverflowError):
        return raw_id.strip()


def _row_to_hubspot_contact(row: Dict[str, str], index: int) -> Dict:
    """
    Convert a mapped row dict into the HubSpot contact JSON shape
    expected by the frontend (HubSpotContact interface).
    """
    contact_id = _clean_id(row.get("id", ""))
    if not contact_id:
        contact_id = str(10000 + index)

    created = row.get("createdate", "")

    return {
        "id": contact_id,
        "properties": {
            "createdate": created,
            "email": row.get("email", ""),
            "firstname": row.get("firstname", ""),
            "lastname": row.get("lastname", ""),
            "hs_object_id": contact_id,
            "lastmodifieddate": created,
            "phone": row.get("phone", ""),
            "hs_lead_status": row.get("hs_lead_status", ""),
            "hubspot_owner_id": row.get("hubspot_owner_id", ""),
            "company": row.get("company", ""),
        },
        "createdAt": created,
        "updatedAt": created,
        "archived": False,
    }


class ExcelContactProvider:
    """
    Loads contacts from a CSV or Excel file and keeps them in memory.
    Supports the same query patterns as the live HubSpot endpoints.

    File resolution order:
      1. Exact path given (CSV or XLSX)
      2. If .xlsx path given, also checks for .csv with same base name
      3. Scans the data directory for any .csv or .xlsx file
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._contacts: List[Dict] = []
        self._loaded = False

    def _resolve_file(self) -> Optional[str]:
        """Find the best available data file."""
        # 1. Exact path
        if os.path.exists(self.file_path):
            return self.file_path

        # 2. Try .csv variant of an .xlsx path (and vice versa)
        base, ext = os.path.splitext(self.file_path)
        alternatives = [base + ".csv", base + ".xlsx", base + ".xls"]
        for alt in alternatives:
            if alt != self.file_path and os.path.exists(alt):
                return alt

        # 3. Scan the data directory for any contacts file
        data_dir = os.path.dirname(self.file_path)
        if os.path.isdir(data_dir):
            for f in sorted(os.listdir(data_dir)):
                if f.startswith("~$"):
                    continue  # skip Excel lock files
                if f.endswith((".csv", ".xlsx", ".xls")):
                    return os.path.join(data_dir, f)

        return None

    # ── Loading ──────────────────────────────────────────────────────────

    def load(self) -> int:
        """Parse the data file and return the number of contacts loaded."""
        resolved = self._resolve_file()
        if not resolved:
            logger.error(f"No contact file found (searched near {self.file_path})")
            return 0

        logger.info(f"Loading contacts from: {resolved}")

        if resolved.endswith(".csv"):
            contacts = self._load_csv(resolved)
        else:
            contacts = self._load_xlsx(resolved)

        self._contacts = contacts
        self._loaded = True
        logger.info(f"Loaded {len(contacts)} contacts from {os.path.basename(resolved)}")
        return len(contacts)

    def _load_csv(self, path: str) -> List[Dict]:
        """Load contacts from a CSV file."""
        contacts = []
        # Try UTF-8 first, fall back to latin-1
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(path, "r", encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    for idx, row in enumerate(reader):
                        mapped = {}
                        for header, value in row.items():
                            field = COLUMN_MAP.get(header)
                            if field:
                                mapped[field] = _normalise_value(value)
                        if mapped.get("firstname") or mapped.get("email"):
                            contacts.append(_row_to_hubspot_contact(mapped, idx))
                return contacts
            except UnicodeDecodeError:
                continue
        logger.error(f"Could not decode CSV file: {path}")
        return []

    def _load_xlsx(self, path: str) -> List[Dict]:
        """Load contacts from an Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=False)
            ws = wb.active
            if ws is None and wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            if ws is None:
                wb.close()
                logger.error(f"No readable sheet in {path}")
                return []

            headers = [cell.value for cell in ws[1]]
            col_map: Dict[int, str] = {}
            for idx, header in enumerate(headers):
                if header in COLUMN_MAP:
                    col_map[idx] = COLUMN_MAP[header]

            contacts = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                mapped = {col_map[i]: _normalise_value(v) for i, v in enumerate(row) if i in col_map}
                if mapped.get("firstname") or mapped.get("email"):
                    contacts.append(_row_to_hubspot_contact(mapped, row_idx))
            wb.close()
            return contacts

        except Exception as e:
            logger.error(f"Failed to read xlsx {path}: {e}")
            return []

    @property
    def contacts(self) -> List[Dict]:
        if not self._loaded:
            self.load()
        return self._contacts

    # ── Query interface (mirrors HubSpot endpoints) ──────────────────────

    def get_all(self) -> List[Dict]:
        return self.contacts

    def get_by_id(self, contact_id: str) -> Optional[Dict]:
        for c in self.contacts:
            if c["id"] == contact_id:
                return c
        return None

    def search(self, query: str) -> List[Dict]:
        q = query.lower()
        return [
            c for c in self.contacts
            if q in (c["properties"].get("firstname") or "").lower()
            or q in (c["properties"].get("lastname") or "").lower()
            or q in (c["properties"].get("email") or "").lower()
            or q in (c["properties"].get("company") or "").lower()
        ]

    def reload(self) -> int:
        """Force re-read from disk (e.g. after file update)."""
        self._loaded = False
        self._contacts = []
        return self.load()
